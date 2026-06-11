from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.utils import timezone
from ..models import Contribucion, Contribuyente

User = get_user_model()


class BaseViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username='testuser', password='testpass123'
        )
        self.contribucion = Contribucion.objects.create(
            obligacion_pago='contribucion',
            numero_identidad='90123456789',
            numero_afiliado='1234567',
            codigo_zpc='ZPC-001',
            periodo_mes=3,
            periodo_anio=2024,
            monto_cup=1500.00,
            tipo_cuenta='natural',
            registrado_por=self.user,
        )
        self.contribuyente = Contribuyente.objects.create(
            carnet_identidad='90123456789',
            numero_contribuyente='1234567',
            codigo_zpc='ZPC-001',
            tipo_cuenta='natural',
        )


class LoginRequiredTest(BaseViewTest):
    def test_dashboard_redirects_if_not_logged_in(self):
        response = self.client.get(reverse('contribuciones:dashboard'))
        self.assertRedirects(response, f'/accounts/login/?next={reverse("contribuciones:dashboard")}')

    def test_lista_redirects_if_not_logged_in(self):
        response = self.client.get(reverse('contribuciones:lista'))
        self.assertRedirects(response, f'/accounts/login/?next={reverse("contribuciones:lista")}')

    def test_crear_redirects_if_not_logged_in(self):
        response = self.client.get(reverse('contribuciones:crear'))
        self.assertRedirects(response, f'/accounts/login/?next={reverse("contribuciones:crear")}')


class DashboardViewTest(BaseViewTest):
    def test_dashboard_status_code(self):
        self.client.login(username='testuser', password='testpass123')
        response = self.client.get(reverse('contribuciones:dashboard'))
        self.assertEqual(response.status_code, 200)

    def test_dashboard_template(self):
        self.client.login(username='testuser', password='testpass123')
        response = self.client.get(reverse('contribuciones:dashboard'))
        self.assertTemplateUsed(response, 'contribuciones/dashboard.html')

    def test_dashboard_context_kpis(self):
        self.client.login(username='testuser', password='testpass123')
        response = self.client.get(reverse('contribuciones:dashboard'))
        self.assertEqual(response.context['total_contribuciones'], 1)
        self.assertEqual(response.context['monto_total'], 1500.00)
        self.assertIn('por_tipo_cuenta', response.context)
        self.assertIn('ultimas_contribuciones', response.context)
        self.assertIn('por_mes', response.context)
        self.assertIn('por_anio', response.context)


class ContribucionListViewTest(BaseViewTest):
    def setUp(self):
        super().setUp()
        self.client.login(username='testuser', password='testpass123')

    def test_lista_status_code(self):
        response = self.client.get(reverse('contribuciones:lista'))
        self.assertEqual(response.status_code, 200)

    def test_lista_template(self):
        response = self.client.get(reverse('contribuciones:lista'))
        self.assertTemplateUsed(response, 'contribuciones/contribucion_list.html')

    def test_lista_contiene_contribucion(self):
        response = self.client.get(reverse('contribuciones:lista'))
        self.assertContains(response, '90123456789')
        self.assertContains(response, '1234567')

    def test_lista_paginacion(self):
        for i in range(20):
            Contribucion.objects.create(
                obligacion_pago='donacion',
                numero_identidad=f'{i:011d}',
                numero_afiliado=f'{i:06d}',
                codigo_zpc='ZPC-001',
                periodo_mes=6,
                periodo_anio=timezone.now().year,
                monto_cup=100.00,
                tipo_cuenta='natural',
                registrado_por=self.user,
            )
        response = self.client.get(reverse('contribuciones:lista'))
        self.assertEqual(len(response.context['contribuciones']), 15)

    def test_lista_busqueda_por_identidad(self):
        response = self.client.get(reverse('contribuciones:lista'), {'q': '90123456789'})
        self.assertEqual(len(response.context['contribuciones']), 1)

    def test_lista_busqueda_sin_resultados(self):
        response = self.client.get(reverse('contribuciones:lista'), {'q': 'NOEXISTE'})
        self.assertEqual(len(response.context['contribuciones']), 0)

    def test_lista_filtro_tipo_cuenta(self):
        response = self.client.get(reverse('contribuciones:lista'), {'tipo_cuenta': 'natural'})
        self.assertEqual(len(response.context['contribuciones']), 1)

    def test_lista_filtro_anio(self):
        response = self.client.get(reverse('contribuciones:lista'), {'anio': 2024})
        self.assertEqual(len(response.context['contribuciones']), 1)

    def test_lista_context_total(self):
        response = self.client.get(reverse('contribuciones:lista'))
        self.assertEqual(response.context['total'], 1)


class ContribucionCreateViewTest(BaseViewTest):
    def setUp(self):
        super().setUp()
        self.client.login(username='testuser', password='testpass123')
        self.valid_data = {
            'obligacion_pago': 'contribucion',
            'numero_identidad': '98765432109',
            'numero_afiliado': '654321',
            'codigo_zpc': 'ZPC-002',
            'periodo_mes': 3,
            'periodo_anio': 2024,
            'monto_cup': 2500.00,
            'tipo_cuenta': 'natural',
        }

    def test_crear_get_status_code(self):
        response = self.client.get(reverse('contribuciones:crear'))
        self.assertEqual(response.status_code, 200)

    def test_crear_template(self):
        response = self.client.get(reverse('contribuciones:crear'))
        self.assertTemplateUsed(response, 'contribuciones/contribucion_form.html')

    def test_crear_context_titulo(self):
        response = self.client.get(reverse('contribuciones:crear'))
        self.assertEqual(response.context['titulo'], 'Nueva Contribución')

    def test_crear_post_valido(self):
        response = self.client.post(reverse('contribuciones:crear'), data=self.valid_data)
        self.assertRedirects(response, reverse('contribuciones:lista'))
        self.assertEqual(Contribucion.objects.count(), 2)

    def test_crear_post_invalido(self):
        data = self.valid_data.copy()
        data['numero_identidad'] = ''
        response = self.client.post(reverse('contribuciones:crear'), data=data)
        self.assertEqual(response.status_code, 200)
        self.assertIn('numero_identidad', response.context['form'].errors)

    def test_crear_asigna_registrado_por(self):
        self.client.post(reverse('contribuciones:crear'), data=self.valid_data)
        c = Contribucion.objects.latest('fecha_registro')
        self.assertEqual(c.registrado_por, self.user)

    def test_crear_con_obligacion_donacion(self):
        data = self.valid_data.copy()
        data['obligacion_pago'] = 'donacion'
        response = self.client.post(reverse('contribuciones:crear'), data=data)
        self.assertRedirects(response, reverse('contribuciones:lista'))
        c = Contribucion.objects.latest('fecha_registro')
        self.assertEqual(c.obligacion_pago, 'donacion')

    def test_crear_form_contiene_select_obligacion(self):
        response = self.client.get(reverse('contribuciones:crear'))
        self.assertContains(response, '<select')
        self.assertContains(response, 'contribucion')
        self.assertContains(response, 'donacion')


class ContribucionDetailViewTest(BaseViewTest):
    def setUp(self):
        super().setUp()
        self.client.login(username='testuser', password='testpass123')

    def test_detalle_status_code(self):
        response = self.client.get(
            reverse('contribuciones:detalle', args=[self.contribucion.pk])
        )
        self.assertEqual(response.status_code, 200)

    def test_detalle_template(self):
        response = self.client.get(
            reverse('contribuciones:detalle', args=[self.contribucion.pk])
        )
        self.assertTemplateUsed(response, 'contribuciones/contribucion_detail.html')

    def test_detalle_context_campos(self):
        response = self.client.get(
            reverse('contribuciones:detalle', args=[self.contribucion.pk])
        )
        campos = response.context['campos']
        labels = [label for label, _ in campos]
        self.assertIn('Obligación de Pago', labels)
        self.assertIn('Monto en CUP', labels)

    def test_detalle_404(self):
        response = self.client.get(
            reverse('contribuciones:detalle', args=[9999])
        )
        self.assertEqual(response.status_code, 404)


class ContribucionUpdateViewTest(BaseViewTest):
    def setUp(self):
        super().setUp()
        self.client.login(username='testuser', password='testpass123')

    def test_editar_get_status_code(self):
        response = self.client.get(
            reverse('contribuciones:editar', args=[self.contribucion.pk])
        )
        self.assertEqual(response.status_code, 200)

    def test_editar_template(self):
        response = self.client.get(
            reverse('contribuciones:editar', args=[self.contribucion.pk])
        )
        self.assertTemplateUsed(response, 'contribuciones/contribucion_form.html')

    def test_editar_post_valido(self):
        response = self.client.post(
            reverse('contribuciones:editar', args=[self.contribucion.pk]),
            data={
                'obligacion_pago': 'donacion',
                'numero_identidad': '90123456789',
                'numero_afiliado': '1234567',
                'codigo_zpc': 'ZPC-001',
                'periodo_mes': 3,
                'periodo_anio': 2024,
                'monto_cup': 3000.00,
                'tipo_cuenta': 'natural',
            }
        )
        self.assertRedirects(response, reverse('contribuciones:lista'))
        self.contribucion.refresh_from_db()
        self.assertEqual(self.contribucion.obligacion_pago, 'donacion')
        self.assertEqual(self.contribucion.monto_cup, 3000.00)

    def test_editar_404(self):
        response = self.client.get(
            reverse('contribuciones:editar', args=[9999])
        )
        self.assertEqual(response.status_code, 404)


class ContribucionDeleteViewTest(BaseViewTest):
    def setUp(self):
        super().setUp()
        self.client.login(username='testuser', password='testpass123')

    def test_eliminar_get_status_code(self):
        response = self.client.get(
            reverse('contribuciones:eliminar', args=[self.contribucion.pk])
        )
        self.assertEqual(response.status_code, 200)

    def test_eliminar_post_redirige(self):
        response = self.client.post(
            reverse('contribuciones:eliminar', args=[self.contribucion.pk])
        )
        self.assertRedirects(response, reverse('contribuciones:lista'))
        self.assertEqual(Contribucion.objects.count(), 0)

    def test_eliminar_404(self):
        response = self.client.post(
            reverse('contribuciones:eliminar', args=[9999])
        )
        self.assertEqual(response.status_code, 404)


class ExportarExcelTest(BaseViewTest):
    def setUp(self):
        super().setUp()
        self.client.login(username='testuser', password='testpass123')

    def test_exportar_excel_status(self):
        response = self.client.get(reverse('contribuciones:exportar_excel'))
        self.assertEqual(response.status_code, 200)

    def test_exportar_excel_content_type(self):
        response = self.client.get(reverse('contribuciones:exportar_excel'))
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_exportar_excel_contiene_datos(self):
        from openpyxl import load_workbook
        from io import BytesIO
        response = self.client.get(reverse('contribuciones:exportar_excel'))
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn('Obligación de Pago', headers)
        self.assertIn('90123456789', [str(ws.cell(row=2, column=c).value) for c in range(1, ws.max_column + 1)])

    def test_exportar_excel_filtro(self):
        from openpyxl import load_workbook
        from io import BytesIO
        response = self.client.get(
            reverse('contribuciones:exportar_excel'), {'q': 'NOEXISTE'}
        )
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.max_row, 1)


class ContribuyenteViewsTest(BaseViewTest):
    def setUp(self):
        super().setUp()
        self.client.login(username='testuser', password='testpass123')

    def test_contribuyente_lista(self):
        response = self.client.get(reverse('contribuciones:contribuyente_lista'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'contribuciones/contribuyente_list.html')

    def test_contribuyente_crear_get(self):
        response = self.client.get(reverse('contribuciones:contribuyente_crear'))
        self.assertEqual(response.status_code, 200)

    def test_contribuyente_crear_post(self):
        response = self.client.post(
            reverse('contribuciones:contribuyente_crear'),
            data={
                'carnet_identidad': '98765432109',
                'numero_contribuyente': '654321',
                'codigo_zpc': 'ZPC-002',
                'tipo_cuenta': 'fiscal',
            }
        )
        self.assertRedirects(response, reverse('contribuciones:contribuyente_lista'))
        self.assertEqual(Contribuyente.objects.count(), 2)

    def test_contribuyente_detalle(self):
        response = self.client.get(
            reverse('contribuciones:contribuyente_detalle', args=[self.contribuyente.pk])
        )
        self.assertEqual(response.status_code, 200)

    def test_contribuyente_editar(self):
        response = self.client.post(
            reverse('contribuciones:contribuyente_editar', args=[self.contribuyente.pk]),
            data={
                'carnet_identidad': '90123456789',
                'numero_contribuyente': '999999',
                'codigo_zpc': 'ZPC-001',
                'tipo_cuenta': 'fiscal',
            }
        )
        self.assertRedirects(response, reverse('contribuciones:contribuyente_lista'))
        self.contribuyente.refresh_from_db()
        self.assertEqual(self.contribuyente.numero_contribuyente, '999999')
        self.assertEqual(self.contribuyente.tipo_cuenta, 'fiscal')

    def test_contribuyente_eliminar(self):
        pk = self.contribuyente.pk
        response = self.client.post(
            reverse('contribuciones:contribuyente_eliminar', args=[pk])
        )
        self.assertRedirects(response, reverse('contribuciones:contribuyente_lista'))
        self.assertEqual(Contribuyente.objects.count(), 0)
