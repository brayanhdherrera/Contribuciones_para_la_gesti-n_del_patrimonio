import csv
import io
import json
import re
import xml.etree.ElementTree as ET
from decimal import Decimal, InvalidOperation
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse_lazy
from django.db.models import Q, Sum, Count
from django.db.models.functions import TruncMonth
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
)
from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Contribucion, Contribuyente, Organismo
from .forms import (
    ContribucionForm, BusquedaForm, ContribuyenteForm,
    ImportarContribucionForm, ImportarContribuyenteForm, ConfirmarImportacionForm,
)


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'contribuciones/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        contribuciones = Contribucion.objects.all()
        contribuyentes = Contribuyente.objects.all()

        # ── KPI principales ──
        ctx['total_contribuciones']  = contribuciones.count()
        ctx['total_contribuyentes']  = contribuyentes.count()
        ctx['monto_total']           = contribuciones.aggregate(t=Sum('monto_cup'))['t'] or 0

        # ── Último Estado de Cuenta importado ──
        from estado_cuenta.models import ArchivoImportado, MovimientoEstadoCuenta
        ultimo_archivo = ArchivoImportado.objects.order_by('-fecha_subida').first()
        ctx['ultimo_estado_cuenta'] = ultimo_archivo
        if ultimo_archivo:
            ctx['ultimo_ec_movimientos'] = ultimo_archivo.movimientos.all()[:5]
            ctx['ultimo_ec_total'] = ultimo_archivo.movimientos.aggregate(
                t=Sum('principal'), i=Sum('impuesto_total')
            )
        else:
            ctx['ultimo_ec_movimientos'] = []
            ctx['ultimo_ec_total'] = {'t': 0, 'i': 0}

        # ── Por tipo de cuenta ──
        ctx['por_tipo_cuenta'] = (
            contribuciones.values('tipo_cuenta')
            .annotate(total=Count('id'), monto=Sum('monto_cup'))
            .order_by('-total')
        )

        # ── Últimas contribuciones ──
        ctx['ultimas_contribuciones'] = contribuciones.select_related('registrado_por').order_by('-fecha_registro')[:10]

        # ── Contribuciones por mes (últimos 12) ──
        ctx['por_mes'] = (
            contribuciones.annotate(mes=TruncMonth('fecha_registro'))
            .values('mes')
            .annotate(total=Count('id'), monto=Sum('monto_cup'))
            .order_by('-mes')[:12]
        )

        # ── Por año ──
        ctx['por_anio'] = (
            contribuciones.values('periodo_anio')
            .annotate(total=Count('id'), monto=Sum('monto_cup'))
            .order_by('-periodo_anio')
        )

        return ctx


class BusquedaMixin:
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['busqueda_form'] = BusquedaForm(self.request.GET or None)
        return ctx


class ContribucionListView(LoginRequiredMixin, BusquedaMixin, ListView):
    model              = Contribucion
    template_name      = 'contribuciones/contribucion_list.html'
    context_object_name = 'contribuciones'
    paginate_by        = 15

    def get_queryset(self):
        qs = super().get_queryset().select_related('registrado_por')
        q           = self.request.GET.get('q', '').strip()
        tipo_cuenta = self.request.GET.get('tipo_cuenta', '').strip()
        anio        = self.request.GET.get('anio', '').strip()

        if q:
            qs = qs.filter(
                Q(nombre__icontains=q)                   |
                Q(numero_identidad__icontains=q)        |
                Q(numero_afiliado__icontains=q) |
                Q(codigo_zpc__icontains=q)               |
                Q(obligacion_pago__icontains=q)
            )
        if tipo_cuenta:
            qs = qs.filter(tipo_cuenta=tipo_cuenta)
        if anio and anio.isdigit():
            qs = qs.filter(periodo_anio=int(anio))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['total'] = self.get_queryset().count()
        ctx['q']     = self.request.GET.get('q', '')
        return ctx


class ContribucionDetailView(LoginRequiredMixin, DetailView):
    model               = Contribucion
    template_name       = 'contribuciones/contribucion_detail.html'
    context_object_name = 'contribucion'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        c   = self.object
        nombre_label = 'Nombre de la Entidad' if c.tipo_cuenta == 'fiscal' else 'Nombre del Contribuyente'
        ctx['campos'] = [
            (nombre_label,                 c.nombre or '—'),
            ('Obligación de Pago',         c.obligacion_pago),
            ('Número de Identidad',         c.numero_identidad),
            ('Número de Afiliado', c.numero_afiliado),
            ('Código ZPC',                  c.codigo_zpc),
            ('Organismo',                   c.organismo or '—'),
            ('Dirección',                   c.direccion or '—'),
            ('Nombre del Establecimiento',  c.nombre_establecimiento or '—'),
            ('Período',                     c.periodo_display),
            ('Monto en CUP',                f'{c.monto_cup} CUP'),
            ('Tipo de Cuenta',              c.get_tipo_cuenta_display()),
            ('Registrado por',              c.registrado_por or '—'),
            ('Fecha de Registro',           c.fecha_registro.strftime('%d/%m/%Y %H:%M')),
            ('Última Modificación',         c.fecha_modificacion.strftime('%d/%m/%Y %H:%M')),
        ]
        return ctx


class ContribucionCreateView(LoginRequiredMixin, CreateView):
    model         = Contribucion
    form_class    = ContribucionForm
    template_name = 'contribuciones/contribucion_form.html'
    success_url   = reverse_lazy('contribuciones:lista')

    def form_valid(self, form):
        form.instance.registrado_por = self.request.user

        ci = form.cleaned_data.get('numero_identidad', '')
        nombre = form.cleaned_data.get('nombre', '')
        if not Contribuyente.objects.filter(carnet_identidad=ci).exists():
            Contribuyente.objects.create(
                nombre=nombre,
                carnet_identidad=ci,
                numero_contribuyente=form.cleaned_data.get('numero_afiliado', ''),
                codigo_zpc=form.cleaned_data.get('codigo_zpc', ''),
                tipo_cuenta=form.cleaned_data.get('tipo_cuenta', 'natural'),
                organismo=form.cleaned_data.get('organismo'),
                direccion=form.cleaned_data.get('direccion', ''),
                nombre_establecimiento=form.cleaned_data.get('nombre_establecimiento', ''),
            )

        messages.success(self.request, '✔ Contribución registrada exitosamente.')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, '✖ Por favor corrija los errores en el formulario.')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Nueva Contribución'
        ctx['accion'] = 'Registrar'
        return ctx


class ContribucionUpdateView(LoginRequiredMixin, UpdateView):
    model         = Contribucion
    form_class    = ContribucionForm
    template_name = 'contribuciones/contribucion_form.html'
    success_url   = reverse_lazy('contribuciones:lista')

    def form_valid(self, form):
        messages.success(
            self.request,
            f'✔ Contribución #{self.object.pk} actualizada exitosamente.'
        )
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, '✖ Por favor corrija los errores en el formulario.')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f'Editar Contribución #{self.object.pk}'
        ctx['accion'] = 'Guardar cambios'
        return ctx


class ContribucionDeleteView(LoginRequiredMixin, DeleteView):
    model               = Contribucion
    template_name       = 'contribuciones/contribucion_confirm_delete.html'
    context_object_name = 'contribucion'
    success_url         = reverse_lazy('contribuciones:lista')

    def form_valid(self, form):
        messages.success(
            self.request,
            f'✔ Contribución #{self.object.pk} eliminada correctamente.'
        )
        return super().form_valid(form)


@login_required
def exportar_excel(request):
    qs = Contribucion.objects.select_related('registrado_por').order_by('-fecha_registro')
    q           = request.GET.get('q', '').strip()
    tipo_cuenta = request.GET.get('tipo_cuenta', '').strip()
    anio        = request.GET.get('anio', '').strip()

    if q:
        qs = qs.filter(
            Q(nombre__icontains=q)                   |
            Q(numero_identidad__icontains=q)        |
            Q(numero_afiliado__icontains=q) |
            Q(codigo_zpc__icontains=q)               |
            Q(obligacion_pago__icontains=q)
        )
    if tipo_cuenta:
        qs = qs.filter(tipo_cuenta=tipo_cuenta)
    if anio and anio.isdigit():
        qs = qs.filter(periodo_anio=int(anio))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Contribuciones'

    headers = [
        'ID', 'Nombre/Entidad', 'Obligación de Pago', 'N° Identidad', 'N° Afiliado',
        'Código ZPC', 'Organismo', 'Dirección', 'Establecimiento',
        'Período', 'Monto CUP', 'Tipo de Cuenta',
        'Registrado por', 'Fecha de Registro',
    ]

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_font = Font(name='Calibri', size=11)
    data_alignment = Alignment(vertical='center')
    monto_alignment = Alignment(horizontal='right', vertical='center')

    for row_idx, c in enumerate(qs, 2):
        row_data = [
            c.pk,
            c.nombre,
            c.obligacion_pago,
            c.numero_identidad,
            c.numero_afiliado,
            c.codigo_zpc,
            str(c.organismo) if c.organismo else '',
            c.direccion,
            c.nombre_establecimiento,
            c.periodo_display,
            float(c.monto_cup),
            c.get_tipo_cuenta_display(),
            c.registrado_por.username if c.registrado_por else '',
            c.fecha_registro.strftime('%d/%m/%Y %H:%M'),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 11:
                cell.alignment = monto_alignment
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = data_alignment

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 22
    ws.column_dimensions['N'].width = 20

    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="contribuciones.xlsx"'
    wb.save(response)
    return response


# ── Autocomplete Contribuyente ──────────────────────────────────────────────────

@login_required
def buscar_contribuyente(request):
    q = request.GET.get('q', '').strip()
    if not q:
        return JsonResponse([], safe=False)

    contribuyentes = Contribuyente.objects.filter(
        Q(carnet_identidad__icontains=q) | Q(numero_contribuyente__icontains=q)
    )[:10]

    data = [{
        'nombre': c.nombre,
        'carnet_identidad': c.carnet_identidad,
        'numero_contribuyente': c.numero_contribuyente,
        'codigo_zpc': c.codigo_zpc,
        'tipo_cuenta': c.tipo_cuenta,
    } for c in contribuyentes]

    return JsonResponse(data, safe=False)


# ── CRUD Contribuyentes ────────────────────────────────────────────────────────

class ContribuyenteListView(LoginRequiredMixin, ListView):
    model              = Contribuyente
    template_name      = 'contribuciones/contribuyente_list.html'
    context_object_name = 'contribuyentes'
    paginate_by        = 15

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(carnet_identidad__icontains=q) |
                Q(numero_contribuyente__icontains=q) |
                Q(codigo_zpc__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['q'] = self.request.GET.get('q', '')
        ctx['total'] = self.get_queryset().count()
        return ctx


class ContribuyenteDetailView(LoginRequiredMixin, DetailView):
    model               = Contribuyente
    template_name       = 'contribuciones/contribuyente_detail.html'
    context_object_name = 'contribuyente'


class ContribuyenteCreateView(LoginRequiredMixin, CreateView):
    model         = Contribuyente
    form_class    = ContribuyenteForm
    template_name = 'contribuciones/contribuyente_form.html'
    success_url   = reverse_lazy('contribuciones:contribuyente_lista')

    def form_valid(self, form):
        messages.success(self.request, '✔ Contribuyente registrado exitosamente.')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, '✖ Por favor corrija los errores en el formulario.')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Nuevo Contribuyente'
        ctx['accion'] = 'Registrar'
        return ctx


class ContribuyenteUpdateView(LoginRequiredMixin, UpdateView):
    model         = Contribuyente
    form_class    = ContribuyenteForm
    template_name = 'contribuciones/contribuyente_form.html'
    success_url   = reverse_lazy('contribuciones:contribuyente_lista')

    def form_valid(self, form):
        messages.success(
            self.request,
            f'✔ Contribuyente #{self.object.pk} actualizado exitosamente.'
        )
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, '✖ Por favor corrija los errores en el formulario.')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f'Editar Contribuyente #{self.object.pk}'
        ctx['accion'] = 'Guardar cambios'
        return ctx


class ContribuyenteDeleteView(LoginRequiredMixin, DeleteView):
    model               = Contribuyente
    template_name       = 'contribuciones/contribuyente_confirm_delete.html'
    context_object_name = 'contribuyente'
    success_url         = reverse_lazy('contribuciones:contribuyente_lista')

    def form_valid(self, form):
        messages.success(
            self.request,
            f'✔ Contribuyente #{self.object.pk} eliminado correctamente.'
        )
        return super().form_valid(form)


# ── Helpers de importación ─────────────────────────────────────────────────

def _detectar_y_decodificar(raw):
    for bom, enc in [(b'\xff\xfe', 'utf-16-le'), (b'\xfe\xff', 'utf-16-be'),
                     (b'\xef\xbb\xbf', 'utf-8-sig')]:
        if raw.startswith(bom):
            return raw.decode(enc)
    for enc in ('utf-8', 'cp1252', 'latin-1'):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode('utf-8', errors='replace')


def _normalizar_header(h):
    return (h.lower().replace(' ', '_').replace('-', '_')
            .replace('.', '').replace('ñ', 'n').replace('ó', 'o')
            .replace('í', 'i').replace('á', 'a').replace('é', 'e')
            .replace('ú', 'u'))


CONTRIBUCION_FIELD_MAP = {
    'nombre': 'nombre',
    'entidad': 'nombre',
    'nombre_del_contribuyente': 'nombre',
    'nombre_de_la_entidad': 'nombre',
    'obligacion_pago': 'obligacion_pago',
    'obligacion': 'obligacion_pago',
    'pago': 'obligacion_pago',
    'numero_identidad': 'numero_identidad',
    'identidad': 'numero_identidad',
    'ci': 'numero_identidad',
    'carnet': 'numero_identidad',
    'carnet_identidad': 'numero_identidad',
    'numero_afiliado': 'numero_afiliado',
    'afiliado': 'numero_afiliado',
    'n_afiliado': 'numero_afiliado',
    'codigo_zpc': 'codigo_zpc',
    'zpc': 'codigo_zpc',
    'organismo': 'organismo',
    'direccion': 'direccion',
    'dirección': 'direccion',
    'nombre_establecimiento': 'nombre_establecimiento',
    'establecimiento': 'nombre_establecimiento',
    'periodo_mes': 'periodo_mes',
    'mes': 'periodo_mes',
    'periodo_anio': 'periodo_anio',
    'anio': 'periodo_anio',
    'año': 'periodo_anio',
    'monto_cup': 'monto_cup',
    'monto': 'monto_cup',
    'cup': 'monto_cup',
    'tipo_cuenta': 'tipo_cuenta',
    'tipo': 'tipo_cuenta',
    'cuenta': 'tipo_cuenta',
}

CONTRIBUYENTE_FIELD_MAP = {
    'nombre': 'nombre',
    'entidad': 'nombre',
    'nombre_del_contribuyente': 'nombre',
    'nombre_de_la_entidad': 'nombre',
    'carnet_identidad': 'carnet_identidad',
    'identidad': 'carnet_identidad',
    'ci': 'carnet_identidad',
    'carnet': 'carnet_identidad',
    'numero_contribuyente': 'numero_contribuyente',
    'n_contribuyente': 'numero_contribuyente',
    'contribuyente': 'numero_contribuyente',
    'codigo_zpc': 'codigo_zpc',
    'zpc': 'codigo_zpc',
    'tipo_cuenta': 'tipo_cuenta',
    'tipo': 'tipo_cuenta',
    'organismo': 'organismo',
    'direccion': 'direccion',
    'dirección': 'direccion',
    'nombre_establecimiento': 'nombre_establecimiento',
    'establecimiento': 'nombre_establecimiento',
}


def _parsear_fila_contribucion(row, organismos_map):
    def v(key):
        return (row.get(key) or '').strip()

    nombre = v('nombre')
    obligacion_pago = v('obligacion_pago')
    numero_identidad = v('numero_identidad')
    numero_afiliado = v('numero_afiliado')
    codigo_zpc = v('codigo_zpc')
    organismo_nombre = v('organismo')
    direccion = v('direccion')
    nombre_establecimiento = v('nombre_establecimiento')
    mes_str = v('periodo_mes')
    anio_str = v('periodo_anio')
    monto_str = v('monto_cup')
    tipo_cuenta = v('tipo_cuenta')

    if not numero_identidad or not numero_afiliado:
        return None

    periodo_mes = None
    try:
        periodo_mes = int(mes_str)
    except (ValueError, TypeError):
        meses_map = {'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
                     'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
                     'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12}
        periodo_mes = meses_map.get(mes_str.strip().lower()) if mes_str else None
    if not periodo_mes or periodo_mes < 1 or periodo_mes > 12:
        periodo_mes = 1

    periodo_anio = None
    try:
        periodo_anio = int(anio_str)
    except (ValueError, TypeError):
        periodo_anio = 2024
    if periodo_anio < 2000:
        periodo_anio = 2024

    monto_cup = Decimal('0')
    try:
        monto_cup = Decimal(monto_str.replace(',', '').replace('$', '').strip() or '0')
    except (InvalidOperation, ValueError, AttributeError):
        pass

    if tipo_cuenta.lower() in ('fiscal', 'juridico', 'jurídico'):
        tipo_cuenta = 'fiscal'
    elif tipo_cuenta.lower() in ('natural', 'personal'):
        tipo_cuenta = 'natural'
    else:
        tipo_cuenta = 'natural'

    organismo = None
    if organismo_nombre:
        org_key = organismo_nombre.strip().lower()
        if org_key in organismos_map:
            organismo = organismos_map[org_key]

    return {
        'nombre': nombre,
        'obligacion_pago': obligacion_pago if obligacion_pago in ('contribucion', 'donacion') else 'contribucion',
        'numero_identidad': re.sub(r'\D', '', numero_identidad)[:11],
        'numero_afiliado': numero_afiliado[:10],
        'codigo_zpc': codigo_zpc.upper()[:20],
        'organismo_nombre': organismo_nombre,
        'organismo': organismo,
        'direccion': direccion[:255],
        'nombre_establecimiento': nombre_establecimiento[:255],
        'periodo_mes': periodo_mes,
        'periodo_anio': periodo_anio,
        'monto_cup': float(monto_cup),
        'tipo_cuenta': tipo_cuenta,
    }


def _parsear_fila_contribuyente(row, organismos_map):
    def v(key):
        return (row.get(key) or '').strip()

    nombre = v('nombre')
    carnet_identidad = v('carnet_identidad')
    numero_contribuyente = v('numero_contribuyente')
    codigo_zpc = v('codigo_zpc')
    organismo_nombre = v('organismo')
    direccion = v('direccion')
    nombre_establecimiento = v('nombre_establecimiento')
    tipo_cuenta = v('tipo_cuenta')

    if not carnet_identidad or not numero_contribuyente:
        return None

    if tipo_cuenta.lower() in ('fiscal', 'juridico', 'jurídico'):
        tipo_cuenta = 'fiscal'
    elif tipo_cuenta.lower() in ('natural', 'personal'):
        tipo_cuenta = 'natural'
    else:
        tipo_cuenta = 'natural'

    organismo = None
    if organismo_nombre:
        org_key = organismo_nombre.strip().lower()
        if org_key in organismos_map:
            organismo = organismos_map[org_key]

    return {
        'nombre': nombre,
        'carnet_identidad': re.sub(r'\D', '', carnet_identidad)[:11],
        'numero_contribuyente': re.sub(r'\D', '', numero_contribuyente)[:20],
        'codigo_zpc': codigo_zpc.upper()[:20],
        'organismo_nombre': organismo_nombre,
        'organismo': organismo,
        'direccion': direccion[:255],
        'nombre_establecimiento': nombre_establecimiento[:255],
        'tipo_cuenta': tipo_cuenta,
    }


def _importar_desde_archivo(archivo, field_map, parse_fila_fn):
    filename = archivo.name.lower()
    raw = archivo.read()
    content = _detectar_y_decodificar(raw)
    lines = [l.strip() for l in content.split('\n') if l.strip()]

    if not lines:
        raise ValueError('El archivo está vacío.')

    if filename.endswith('.xml'):
        return _importar_desde_xml(content, parse_fila_fn)

    separador = ','
    for sep in ('|', '\t', ';', ','):
        if sep in lines[0]:
            separador = sep
            break

    reader = csv.DictReader(io.StringIO(content), delimiter=separador)
    if not reader.fieldnames:
        raise ValueError('No se pudo detectar la cabecera en el archivo.')

    columnas = {}
    for fn in reader.fieldnames:
        norm = _normalizar_header(fn)
        mapped = field_map.get(norm)
        if mapped:
            columnas[fn] = mapped

    organismos_map = {o.nombre.lower(): o for o in Organismo.objects.all()}

    datos = []
    for row in reader:
        row_mapped = {}
        for col_orig, campo in columnas.items():
            val = (row.get(col_orig) or '').strip()
            if val:
                row_mapped[campo] = val
        parsed = parse_fila_fn(row_mapped, organismos_map)
        if parsed:
            datos.append(parsed)

    return datos


def _importar_desde_xml(content, parse_fila_fn):
    parser = ET.XMLParser()
    try:
        root = ET.fromstring(content.encode('utf-8'), parser)
    except ET.ParseError:
        try:
            root = ET.fromstring(content, parser)
        except ET.ParseError as e:
            raise ValueError(f'Error al parsear XML: {e}')

    rows = []
    for child in root:
        if len(child) > 0:
            rows.append(child)

    organismos_map = {o.nombre.lower(): o for o in Organismo.objects.all()}

    datos = []
    for row_elem in rows:
        row_raw = {}
        for field in row_elem:
            tag = _normalizar_header(field.tag)
            text = (field.text or '').strip()
            if text:
                row_raw[tag] = text
        parsed = parse_fila_fn(row_raw, organismos_map)
        if parsed:
            datos.append(parsed)

    return datos


# ── Importar Contribuciones (un solo paso: subir → guardar en BD → mostrar) ─

class ImportarContribucionView(LoginRequiredMixin, TemplateView):
    template_name = 'contribuciones/contribucion_import.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form'] = ImportarContribucionForm()
        return ctx

    def post(self, request, *args, **kwargs):
        form = ImportarContribucionForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})

        archivo = request.FILES['archivo']
        filename = archivo.name

        try:
            datos = _importar_desde_archivo(
                archivo, CONTRIBUCION_FIELD_MAP, _parsear_fila_contribucion
            )
        except ValueError as e:
            messages.error(request, f'✖ {e}')
            return render(request, self.template_name, {'form': ImportarContribucionForm()})

        if not datos:
            messages.warning(request, '⚠ No se encontraron datos válidos en el archivo.')
            return render(request, self.template_name, {'form': ImportarContribucionForm()})

        creados = errores = 0
        registros_creados = []
        for fila in datos:
            try:
                obj = Contribucion.objects.create(
                    nombre=fila.get('nombre', ''),
                    obligacion_pago=fila.get('obligacion_pago', 'contribucion'),
                    numero_identidad=fila.get('numero_identidad', ''),
                    numero_afiliado=fila.get('numero_afiliado', ''),
                    codigo_zpc=fila.get('codigo_zpc', ''),
                    organismo_id=fila['organismo'].pk if fila.get('organismo') else None,
                    direccion=fila.get('direccion', ''),
                    nombre_establecimiento=fila.get('nombre_establecimiento', ''),
                    periodo_mes=fila.get('periodo_mes', 1),
                    periodo_anio=fila.get('periodo_anio', 2024),
                    monto_cup=fila.get('monto_cup', 0),
                    tipo_cuenta=fila.get('tipo_cuenta', 'natural'),
                    registrado_por=request.user,
                )
                creados += 1
                registros_creados.append(fila)
            except Exception:
                errores += 1

        total_monto = sum(float(d.get('monto_cup', 0)) for d in registros_creados)

        messages.success(
            request,
            f'✔ Importación completada: {creados} contribuciones guardadas '
            f'directamente en la base de datos.'
            + (f' {errores} filas omitidas por errores.' if errores else '')
        )

        return render(request, self.template_name, {
            'form': ImportarContribucionForm(),
            'datos': registros_creados,
            'filename': filename,
            'total_datos': creados,
            'total_monto': total_monto,
            'total_errores': errores,
            'show_results': True,
        })


# ── Importar Contribuyentes (un solo paso: subir → guardar en BD → mostrar) ─

class ImportarContribuyenteView(LoginRequiredMixin, TemplateView):
    template_name = 'contribuciones/contribuyente_import.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form'] = ImportarContribuyenteForm()
        return ctx

    def post(self, request, *args, **kwargs):
        form = ImportarContribuyenteForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})

        archivo = request.FILES['archivo']
        filename = archivo.name

        try:
            datos = _importar_desde_archivo(
                archivo, CONTRIBUYENTE_FIELD_MAP, _parsear_fila_contribuyente
            )
        except ValueError as e:
            messages.error(request, f'✖ {e}')
            return render(request, self.template_name, {'form': ImportarContribuyenteForm()})

        if not datos:
            messages.warning(request, '⚠ No se encontraron datos válidos en el archivo.')
            return render(request, self.template_name, {'form': ImportarContribuyenteForm()})

        creados = errores = 0
        registros_creados = []
        for fila in datos:
            try:
                Contribuyente.objects.create(
                    nombre=fila.get('nombre', ''),
                    carnet_identidad=fila.get('carnet_identidad', ''),
                    numero_contribuyente=fila.get('numero_contribuyente', ''),
                    codigo_zpc=fila.get('codigo_zpc', ''),
                    tipo_cuenta=fila.get('tipo_cuenta', 'natural'),
                    organismo_id=fila['organismo'].pk if fila.get('organismo') else None,
                    direccion=fila.get('direccion', ''),
                    nombre_establecimiento=fila.get('nombre_establecimiento', ''),
                )
                creados += 1
                registros_creados.append(fila)
            except Exception:
                errores += 1

        messages.success(
            request,
            f'✔ Importación completada: {creados} contribuyentes guardados '
            f'directamente en la base de datos.'
            + (f' {errores} filas omitidas por errores.' if errores else '')
        )

        return render(request, self.template_name, {
            'form': ImportarContribuyenteForm(),
            'datos': registros_creados,
            'filename': filename,
            'total_datos': creados,
            'total_errores': errores,
            'show_results': True,
        })
